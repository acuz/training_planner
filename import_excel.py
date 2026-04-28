"""
import_excel.py — Parse Angelo Cuzzola.xlsx training sessions into JSON.

Layout (confirmed by inspection):
  Each sheet has up to 4 week blocks.
  Each week block: rows offset from a Monday anchor.
    - Week header at rows 1, 54, 107, 160  (C='Settimana:', D=week_num, E=year)
    - Data header rows (3/4, 56/57, 109/110, 162/163) — skipped
    - Day blocks: each day has 5 consecutive rows
        day-2 row: C='Nuoto'
        day-1 row: C='Bici'
        day   row: C='Rulli'  (A=day_letter, B=date)
        day+1 row: C='Corsa'
        day+2 row: C='Palestra'

  Column mapping (1-indexed):
    A(1)  = day letter (L/M/M/G/V/S/D) — only on Rulli row
    B(2)  = date (datetime) — only on Rulli row, computed by data_only
    C(3)  = discipline name
    D(4)  = distanceMeters (numeric, already in meters — field label says 'metri')
    E(5)  = trainer instructions part 1 (planned description/notes)
    F(6)  = trainer instructions part 2 (continuation)
    L(12) = startTime (datetime — only time part matters)
    M(13) = plannedDuration (time or datetime — extract time part)
    O(15) = notes (can overflow across all 5 rows in a day block)
    U(21) = weightKg
    V(22) = RPE (integer)

  Cell color on col C → intensity:
    FF00FA00 = green  → "low"
    FFFFFF00 = yellow → "medium"
    FFFFC000 = orange → "hard"
    FFFF0000 = red    → "high"
    no color          → null

  Notes handling:
    - The E+F columns = planned trainer instructions for each discipline
    - The O column note often overflows across multiple rows in the same day block.
      We concatenate all O values in the 5-row day block and attribute them to
      the discipline(s) that have actual data (D, L, or M non-null).
      If only one discipline in the block has data → assign all O text to it.
      If multiple disciplines have data → assign O text only to the row that has
      O directly on it; overflow rows get concatenated to the first discipline's note.

  Week ID: ISO 8601 week of the Monday date.
  Date per discipline: same as the day's date (from Rulli row B column).
"""

import openpyxl
import json
import datetime
import re
import os

# ── paths ──────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH  = os.path.join(SCRIPT_DIR, "Angelo Cuzzola.xlsx")
JSON_OUT   = os.path.join(SCRIPT_DIR, "sessions_import.json")

# ── color → intensity ───────────────────────────────────────────────────────
COLOR_INTENSITY = {
    "FF00FA00": "low",
    "FFFFFF00": "medium",
    "FFFFC000": "hard",
    "FFFF0000": "high",
}

# ── disciplines in order (relative row offset from the Rulli/day row) ───────
# Offset: Nuoto=-2, Bici=-1, Rulli=0, Corsa=+1, Palestra=+2
DISCIPLINE_OFFSETS = {
    "Nuoto":   -2,
    "Bici":    -1,
    "Rulli":    0,
    "Corsa":   +1,
    "Palestra": +2,
}

# ── week block anchors per sheet (row of 'Settimana:' header in col C) ──────
# These are fixed structural positions confirmed by inspection.
WEEK_ANCHORS = {
    # (header_row, first_day_rulli_row)
    # header_row has C='Settimana:', D=week_num, E=year
    # day rows (Rulli row) within the block are at: first_rulli + 0*5, +1*5, ..., +6*5
    # But the first day block starts 2 rows after the header's discipline labels.
    # Confirmed positions:
    "20 apr - 17 mag": [
        {"header_row": 1,   "first_rulli_row": 7},
        {"header_row": 54,  "first_rulli_row": 60},
        {"header_row": 107, "first_rulli_row": 113},
        {"header_row": 160, "first_rulli_row": 166},
    ],
    "pre...": [
        {"header_row": 1,   "first_rulli_row": 7},
        {"header_row": 54,  "first_rulli_row": 60},
        {"header_row": 107, "first_rulli_row": 113},
        {"header_row": 160, "first_rulli_row": 166},
    ],
}

# ── helpers ──────────────────────────────────────────────────────────────────

def get_cell_intensity(ws_raw, row, col=3):
    """Return intensity string based on col C fill color, or None."""
    cell = ws_raw.cell(row=row, column=col)
    fg = cell.fill.fgColor
    rgb = getattr(fg, "rgb", None)
    return COLOR_INTENSITY.get(rgb)


def to_time_str(val):
    """
    Convert openpyxl time/datetime value to HH:MM:SS string.
    - datetime.time  → direct
    - datetime.datetime (epoch 1900) → extract time portion
    Returns None if val is None or not a time.
    """
    if val is None:
        return None
    if isinstance(val, datetime.time):
        return val.strftime("%H:%M:%S")
    if isinstance(val, datetime.datetime):
        return val.strftime("%H:%M:%S")
    return None


def to_hhmm_str(val):
    """Convert time value to HH:MM string for startTime."""
    s = to_time_str(val)
    if s:
        return s[:5]  # HH:MM
    return None


def to_distance(val):
    """Convert distance cell value to int meters, or None."""
    if val is None:
        return None
    try:
        d = int(float(val))
        return d if d > 0 else None
    except (TypeError, ValueError):
        return None


def iso_week_id(dt):
    """Return ISO week string like '2026-W17' from a date/datetime."""
    iso = dt.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def collect_o_text(ws, rows):
    """Concatenate all non-None O column (col 15) values from a list of rows."""
    parts = []
    for r in rows:
        val = ws.cell(row=r, column=15).value
        if val and str(val).strip():
            parts.append(str(val).strip())
    return " ".join(parts) if parts else None


def collect_ef_text(ws, row):
    """Collect trainer instruction text from E (col 5) and F (col 6)."""
    parts = []
    for col in (5, 6):
        val = ws.cell(row=row, column=col).value
        if val and str(val).strip():
            parts.append(str(val).strip())
    return " ".join(parts) if parts else None


def parse_week_block(ws_data, ws_raw, first_rulli_row):
    """
    Parse one week block (7 days × 5 disciplines).
    Returns list of session dicts.
    """
    sessions = []

    for day_offset in range(7):  # 0=Monday .. 6=Sunday
        rulli_row = first_rulli_row + day_offset * 5

        # Read date from B column (data_only gives computed value)
        date_cell = ws_data.cell(row=rulli_row, column=2)
        date_val = date_cell.value

        if not isinstance(date_val, (datetime.datetime, datetime.date)):
            # No date computed (template rows, etc.) — skip this day
            continue

        if isinstance(date_val, datetime.datetime):
            day_date = date_val.date()
        else:
            day_date = date_val

        date_str  = day_date.strftime("%Y-%m-%d")
        week_id   = iso_week_id(day_date)

        # The 5 rows in this day block
        block_rows = list(range(rulli_row - 2, rulli_row + 3))

        # Collect O text for the entire block (notes overflow)
        all_o_text = collect_o_text(ws_data, block_rows)

        # Find which rows have actual data
        discipline_order = ["Nuoto", "Bici", "Rulli", "Corsa", "Palestra"]
        disc_rows = {
            "Nuoto":    rulli_row - 2,
            "Bici":     rulli_row - 1,
            "Rulli":    rulli_row,
            "Corsa":    rulli_row + 1,
            "Palestra": rulli_row + 2,
        }

        # For each discipline, check if it has data
        rows_with_data = []
        for disc in discipline_order:
            r = disc_rows[disc]
            d_val = ws_data.cell(row=r, column=4).value   # distance
            l_val = ws_data.cell(row=r, column=12).value  # startTime
            m_val = ws_data.cell(row=r, column=13).value  # duration
            e_val = ws_data.cell(row=r, column=5).value   # trainer instructions
            if any(v is not None for v in [d_val, l_val, m_val]):
                rows_with_data.append(disc)

        # Also check rows that have only O notes on their own row (no E/D/L/M)
        # Those are likely overflow notes — we don't create sessions for those.

        # Note assignment: if only 1 discipline has data, it gets all O text.
        # If 0 disciplines have data but there's an O note, skip (just a comment).
        # If multiple disciplines have data, each gets its own O row text.

        for disc in discipline_order:
            r = disc_rows[disc]

            d_val = ws_data.cell(row=r, column=4).value
            l_val = ws_data.cell(row=r, column=12).value
            m_val = ws_data.cell(row=r, column=13).value
            e_val = ws_data.cell(row=r, column=5).value
            u_val = ws_data.cell(row=r, column=21).value  # weightKg
            v_val = ws_data.cell(row=r, column=22).value  # RPE

            has_data = any(v is not None for v in [d_val, l_val, m_val])

            if not has_data:
                continue

            # Intensity from color of col C
            intensity = get_cell_intensity(ws_raw, r, col=3)

            # Trainer notes: E + F columns on this row
            trainer_notes_ef = collect_ef_text(ws_data, r)

            # Athlete/session notes from O column
            # If only one discipline has data, give it all O text from the block
            # Otherwise give only the O text on its own row
            if len(rows_with_data) == 1:
                o_notes = all_o_text
            else:
                o_notes = ws_data.cell(row=r, column=15).value
                o_notes = str(o_notes).strip() if o_notes else None

            # Note classification:
            # - E+F columns = planned trainer instructions → always trainerNotes
            # - O column = context that is either trainer explanation (planned sessions)
            #              or athlete notes (completed sessions).
            # A session is "completed" if the athlete has filled in startTime or RPE.
            # For completed sessions: O → athleteNotes.
            # For planned sessions (no startTime, no RPE): O → append to trainerNotes.
            is_completed = (l_val is not None) or (v_val is not None)

            if is_completed:
                trainer_notes = trainer_notes_ef if trainer_notes_ef else None
                athlete_notes = o_notes if o_notes else None
            else:
                # Merge E+F and O all into trainerNotes
                parts = [p for p in [trainer_notes_ef, o_notes] if p]
                trainer_notes = " ".join(parts) if parts else None
                athlete_notes = None

            # Distance
            distance = to_distance(d_val)

            # Start time
            start_time = None
            if l_val is not None and isinstance(l_val, (datetime.datetime, datetime.time)):
                start_time = to_hhmm_str(l_val)

            # Planned duration
            planned_duration = to_time_str(m_val)

            # Weight and RPE
            weight_kg = float(u_val) if u_val is not None else None
            rpe = int(v_val) if v_val is not None else None

            session = {
                "weekId":          week_id,
                "date":            date_str,
                "discipline":      disc,
                "distanceMeters":  distance,
                "startTime":       start_time,
                "plannedDuration": planned_duration,
                "trainerNotes":    trainer_notes,
                "intensity":       intensity,
                "actualDuration":  None,   # athlete fills
                "athleteNotes":    athlete_notes,
                "rpe":             rpe,
                "weightKg":        weight_kg,
            }
            sessions.append(session)

    return sessions


def parse_sheet(wb_data, wb_raw, sheet_name):
    """Parse all week blocks in a sheet."""
    if sheet_name not in wb_data.sheetnames:
        print(f"  Sheet '{sheet_name}' not found, skipping.")
        return []

    ws_data = wb_data[sheet_name]
    ws_raw  = wb_raw[sheet_name]

    anchors = WEEK_ANCHORS.get(sheet_name, [])
    if not anchors:
        print(f"  No anchors defined for '{sheet_name}', skipping.")
        return []

    all_sessions = []
    for anchor in anchors:
        first_rulli = anchor["first_rulli_row"]
        sessions = parse_week_block(ws_data, ws_raw, first_rulli)
        all_sessions.extend(sessions)
        print(f"  Block starting rulli row {first_rulli}: {len(sessions)} sessions")

    return all_sessions


def main():
    import shutil, tempfile

    # The file may be locked by OneDrive/Excel; copy to a temp file first.
    print(f"Loading: {XLSX_PATH}")
    tmp_path = os.path.join(tempfile.gettempdir(), "Angelo_Cuzzola_tmp.xlsx")
    shutil.copy2(XLSX_PATH, tmp_path)

    wb_data = openpyxl.load_workbook(tmp_path, data_only=True)
    wb_raw  = openpyxl.load_workbook(tmp_path, data_only=False)

    all_sessions = []

    sheets_to_parse = ["pre...", "20 apr - 17 mag"]
    for sheet_name in sheets_to_parse:
        print(f"\nParsing sheet: '{sheet_name}'")
        sessions = parse_sheet(wb_data, wb_raw, sheet_name)
        all_sessions.extend(sessions)

    # Sort by date then discipline
    discipline_order_index = {d: i for i, d in enumerate(["Nuoto", "Bici", "Rulli", "Corsa", "Palestra"])}
    all_sessions.sort(key=lambda s: (s["date"], discipline_order_index.get(s["discipline"], 99)))

    print(f"\nTotal sessions found: {len(all_sessions)}")

    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(all_sessions, f, ensure_ascii=False, indent=2)

    print(f"Written to: {JSON_OUT}")
    print()
    print("=== First 5 sessions ===")
    for s in all_sessions[:5]:
        print(json.dumps(s, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
